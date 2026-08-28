#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import hashlib
import html
import io
import json
import logging
import math
import os
import re
import shutil
import subprocess
import sys
import tempfile
import time
import traceback
import urllib.parse
import warnings
import zipfile
from collections import defaultdict
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Iterator, Mapping, Sequence

import geopandas as gpd
import pandas as pd
import requests
import urllib3
from bs4 import BeautifulSoup
from lxml import etree
from openpyxl import load_workbook
from pyproj import CRS
from requests.adapters import HTTPAdapter
from shapely.geometry import GeometryCollection, LineString, MultiLineString, MultiPoint, MultiPolygon, Point, Polygon, mapping, shape
from shapely.ops import transform
from urllib3.util.retry import Retry

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
warnings.filterwarnings("ignore", category=UserWarning)

KML_NS = "http://www.opengis.net/kml/2.2"
GX_NS = "http://www.google.com/kml/ext/2.2"
K = f"{{{KML_NS}}}"
NSMAP = {None: KML_NS, "gx": GX_NS}
LOG = logging.getLogger("deep-global-marine")


def utcnow() -> str:
    return datetime.now(timezone.utc).isoformat()


def clean_text(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, (dict, list, tuple)):
        try:
            s = json.dumps(v, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
        except Exception:
            s = str(v)
    elif isinstance(v, float):
        if not math.isfinite(v):
            return ""
        s = f"{v:.12g}"
    else:
        s = str(v)
    return "".join(ch for ch in s if ch in "\t\n\r" or ord(ch) >= 0x20)


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def sub(parent: etree._Element, tag: str, text: Any | None = None, **attrs: str) -> etree._Element:
    e = etree.SubElement(parent, K + tag, **attrs)
    if text is not None:
        e.text = clean_text(text)
    return e


def safe_id(s: str) -> str:
    x = re.sub(r"[^A-Za-z0-9_.:-]+", "_", s).strip("_")
    if not x or not re.match(r"[A-Za-z_]", x):
        x = "id_" + x
    return x[:220]


def field_name(s: Any) -> str:
    return safe_id(str(s))


@dataclass
class SourceResult:
    source_id: str
    region: str
    country: str
    title: str
    category: str
    authority: str
    source_url: str
    method: str
    status: str = "pending"
    retrieved_at: str = ""
    downloaded_bytes: int = 0
    sha256: str = ""
    feature_count: int = 0
    vertex_count: int = 0
    geometry_types: dict[str, int] = field(default_factory=dict)
    route_version: str = ""
    authority_class: str = "statutory_or_official"
    datum: str = "WGS84/EPSG:4326"
    coverage: str = ""
    notes: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


class SessionFactory:
    @staticmethod
    def make(headers: Mapping[str, str] | None = None) -> requests.Session:
        s = requests.Session()
        retry = Retry(total=5, connect=5, read=5, status=5, backoff_factor=1.1,
            status_forcelist=(408, 425, 429, 500, 502, 503, 504),
            allowed_methods=frozenset(["GET", "POST", "HEAD"]), respect_retry_after_header=True)
        s.mount("https://", HTTPAdapter(max_retries=retry, pool_connections=20, pool_maxsize=20))
        s.mount("http://", HTTPAdapter(max_retries=retry, pool_connections=20, pool_maxsize=20))
        s.headers.update({"User-Agent": "GlobalMarinePermitGeometry/2026.08 (+public-research; static snapshot)", "Accept": "*/*"})
        if headers: s.headers.update(headers)
        return s


class KMLBuilder:
    def __init__(self, title: str):
        self.root = etree.Element(K + "kml", nsmap=NSMAP)
        self.doc = sub(self.root, "Document")
        sub(self.doc, "name", title); sub(self.doc, "open", "0")
        sub(self.doc, "description", "Fully embedded deep-source supplement built from downloadable official GIS, permit/gazette coordinates and official navigation notices. No NetworkLinks or runtime APIs. Historical temporary work notices are segregated and date-labelled.")
        self.region_folders: dict[str, etree._Element] = {}
        self._add_styles()

    def _add_styles(self) -> None:
        styles = {"cable_route": ("ffff00ff", "2.4", "2200ffff"), "cable_zone": ("ffff00ff", "2.0", "3300ffff"), "fish_closure": ("ff0000ff", "2.0", "330000ff"), "fish_management": ("ff00a5ff", "1.8", "2200a5ff"), "marine_protected": ("ff00aa00", "1.8", "3300aa00"), "permit_work": ("ff00ffff", "2.0", "3300ffff"), "infrastructure": ("ffffaa00", "2.0", "22ffaa00"), "context": ("ffff0000", "1.5", "110000ff"), "failed": ("ff888888", "1.0", "11888888")}
        for sid, (line, width, poly) in styles.items():
            st = sub(self.doc, "Style", id=sid); ls = sub(st, "LineStyle"); sub(ls, "color", line); sub(ls, "width", width); ps = sub(st, "PolyStyle"); sub(ps, "color", poly); sub(ps, "outline", "1"); isty = sub(st, "IconStyle"); sub(isty, "color", line); sub(isty, "scale", "0.8")

    def folder(self, region: str, title: str, description: str = "") -> etree._Element:
        rf = self.region_folders.get(region)
        if rf is None:
            rf = sub(self.doc, "Folder"); sub(rf, "name", region); sub(rf, "open", "0"); sub(rf, "visibility", "0"); self.region_folders[region] = rf
        f = sub(rf, "Folder"); sub(f, "name", title); sub(f, "open", "0"); sub(f, "visibility", "0")
        if description: sub(f, "description", description)
        return f

    def add_geodataframe(self, folder: etree._Element, gdf: gpd.GeoDataFrame, result: SourceResult, style: str) -> None:
        if gdf.empty: return
        if gdf.crs is None:
            result.notes.append("Source CRS missing; treated as EPSG:4326 only after coordinate-range validation"); gdf = gdf.set_crs(4326, allow_override=True)
        else:
            try: gdf = gdf.to_crs(4326)
            except Exception as exc: result.notes.append(f"CRS transform failed ({gdf.crs}); attempted source coordinates: {exc}")
        geom_counts: defaultdict[str, int] = defaultdict(int)
        for idx, row in gdf.iterrows():
            geom = row.geometry
            if geom is None or geom.is_empty: continue
            props = {str(k): v for k, v in row.items() if k != gdf.geometry.name}
            pm = sub(folder, "Placemark"); sub(pm, "name", choose_name(props, f"{result.title} #{result.feature_count + 1}")); sub(pm, "styleUrl", f"#{style}"); self._add_extended(pm, props, result)
            try: gt, vc = self._add_shape(pm, geom)
            except Exception as exc:
                folder.remove(pm); result.notes.append(f"Skipped invalid geometry at source row {idx}: {exc}"); continue
            result.feature_count += 1; result.vertex_count += vc; geom_counts[gt] += 1
        for k, v in geom_counts.items(): result.geometry_types[k] = result.geometry_types.get(k, 0) + v

    def add_geojson(self, folder: etree._Element, obj: Mapping[str, Any], result: SourceResult, style: str) -> None:
        feats = obj.get("features") or [] if obj.get("type") == "FeatureCollection" else [obj] if obj.get("type") == "Feature" else [{"type": "Feature", "properties": {}, "geometry": obj}]
        rows=[]; geoms=[]
        for f in feats:
            gj=f.get("geometry")
            if not gj: continue
            try: geoms.append(shape(gj)); rows.append(f.get("properties") or {})
            except Exception as exc: result.notes.append(f"Invalid GeoJSON feature skipped: {exc}")
        if geoms: self.add_geodataframe(folder, gpd.GeoDataFrame(rows, geometry=geoms, crs=4326), result, style)

    def add_manual(self, folder: etree._Element, geom: Any, props: Mapping[str, Any], result: SourceResult, style: str) -> None:
        self.add_geodataframe(folder, gpd.GeoDataFrame([dict(props)], geometry=[geom], crs=4326), result, style)

    def _add_extended(self, pm: etree._Element, props: Mapping[str, Any], result: SourceResult) -> None:
        ed = sub(pm, "ExtendedData")
        merged={"source_id":result.source_id,"source_title":result.title,"authority":result.authority,"authority_class":result.authority_class,"source_url":result.source_url,"retrieved_at":result.retrieved_at,"category":result.category,"route_version":result.route_version,"datum":result.datum,"coverage":result.coverage}
        for k,v in props.items():
            key=field_name(k)
            if key in merged: key="source_"+key
            merged[key]=v
        for key,val in merged.items(): d=sub(ed,"Data",name=field_name(key)); sub(d,"value",clean_text(val))

    def _add_shape(self, parent: etree._Element, geom: Any) -> tuple[str,int]:
        if isinstance(geom,Point): g=sub(parent,"Point"); sub(g,"coordinates",coord(geom.x,geom.y)); return "Point",1
        if isinstance(geom,MultiPoint):
            mg=sub(parent,"MultiGeometry")
            for p in geom.geoms: g=sub(mg,"Point"); sub(g,"coordinates",coord(p.x,p.y))
            return "MultiPoint",len(geom.geoms)
        if isinstance(geom,LineString): g=sub(parent,"LineString"); sub(g,"tessellate","1"); sub(g,"coordinates"," ".join(coord(x,y) for x,y,*_ in geom.coords)); return "LineString",len(geom.coords)
        if isinstance(geom,MultiLineString):
            mg=sub(parent,"MultiGeometry"); n=0
            for line in geom.geoms: g=sub(mg,"LineString"); sub(g,"tessellate","1"); sub(g,"coordinates"," ".join(coord(x,y) for x,y,*_ in line.coords)); n+=len(line.coords)
            return "MultiLineString",n
        if isinstance(geom,Polygon): return self._add_polygon(parent,geom),polygon_vertices(geom)
        if isinstance(geom,MultiPolygon):
            mg=sub(parent,"MultiGeometry"); n=0
            for p in geom.geoms: self._add_polygon(mg,p); n+=polygon_vertices(p)
            return "MultiPolygon",n
        if isinstance(geom,GeometryCollection):
            mg=sub(parent,"MultiGeometry"); n=0
            for g0 in geom.geoms: _,c=self._add_shape(mg,g0); n+=c
            return "GeometryCollection",n
        raise TypeError(type(geom).__name__)

    def _add_polygon(self,parent:etree._Element,p:Polygon)->str:
        poly=sub(parent,"Polygon"); sub(poly,"tessellate","1"); ob=sub(poly,"outerBoundaryIs"); lr=sub(ob,"LinearRing"); sub(lr,"coordinates"," ".join(coord(x,y) for x,y,*_ in p.exterior.coords))
        for ring in p.interiors: ib=sub(poly,"innerBoundaryIs"); lr=sub(ib,"LinearRing"); sub(lr,"coordinates"," ".join(coord(x,y) for x,y,*_ in ring.coords))
        return "Polygon"

    def write(self,out_kml:Path,out_kmz:Path)->None:
        etree.ElementTree(self.root).write(str(out_kml),encoding="UTF-8",xml_declaration=True,pretty_print=False)
        with zipfile.ZipFile(out_kmz,"w",compression=zipfile.ZIP_DEFLATED,compresslevel=9,allowZip64=True) as z: z.write(out_kml,"doc.kml")


def coord(x:float,y:float)->str:
    x=float(x); y=float(y)
    if not (-180.000001<=x<=180.000001 and -90.000001<=y<=90.000001): raise ValueError(f"coordinate out of range {x},{y}")
    return f"{x:.7f},{y:.7f},0"

def polygon_vertices(p:Polygon)->int: return len(p.exterior.coords)+sum(len(r.coords) for r in p.interiors)

def choose_name(props:Mapping[str,Any],fallback:str)->str:
    for k in ["name","Name","NAME","title","Title","TITLE","area","Area","AREA","label","LABEL","description","DESCRIPTIO","OBJECTID","FID","id"]:
        v=props.get(k)
        if v not in (None,"","nan"): return clean_text(v)[:350]
    for k,v in props.items():
        if re.search(r"name|title|area|zone|site|project|closure",str(k),re.I) and v not in (None,"","nan"): return clean_text(v)[:350]
    return fallback

def download(session:requests.Session,url:str,path:Path,timeout:int=300,verify:bool=True)->tuple[bytes,requests.Response]:
    LOG.info("GET %s",url); r=session.get(url,timeout=timeout,allow_redirects=True,verify=verify); r.raise_for_status(); path.parent.mkdir(parents=True,exist_ok=True); path.write_bytes(r.content); return r.content,r

def read_vector_archive(path:Path)->list[tuple[str,gpd.GeoDataFrame]]:
    out=[]
    with tempfile.TemporaryDirectory() as td:
        root=Path(td)
        if zipfile.is_zipfile(path):
            with zipfile.ZipFile(path) as z:z.extractall(root)
        else: shutil.copy2(path,root/path.name)
        files=[]
        for ext in ("*.shp","*.geojson","*.json","*.gpkg","*.kml"): files.extend(root.rglob(ext))
        seen=set()
        for f in files:
            if f.suffix.lower()==".json":
                try:
                    obj=json.loads(f.read_text(encoding="utf-8-sig"))
                    if not isinstance(obj,dict) or obj.get("type") not in {"FeatureCollection","Feature"}:continue
                except Exception:continue
            if str(f.resolve()) in seen:continue
            seen.add(str(f.resolve()))
            try:gdf=gpd.read_file(f)
            except Exception as exc:LOG.warning("Could not read %s: %s",f,exc);continue
            if not gdf.empty:out.append((f.stem,gdf))
    return out

def process_zip(builder:KMLBuilder,session:requests.Session,spec:Mapping[str,Any],work:Path)->SourceResult:
    r=SourceResult(**{k:spec[k] for k in ("source_id","region","country","title","category","authority","source_url","method")});r.retrieved_at=utcnow();r.route_version=spec.get("route_version","official current snapshot");r.coverage=spec.get("coverage","")
    try:
        path=work/f"{r.source_id}.zip";data,_=download(session,r.source_url,path,verify=spec.get("verify",True));r.downloaded_bytes=len(data);r.sha256=sha256_bytes(data);layers=read_vector_archive(path)
        if not layers:raise RuntimeError("archive contained no readable vector layers")
        top=builder.folder(r.region,r.title,f"{r.authority}\n{r.source_url}\nRetrieved {r.retrieved_at}")
        for layer_name,gdf in layers:lf=sub(top,"Folder");sub(lf,"name",layer_name);sub(lf,"visibility","0");builder.add_geodataframe(lf,gdf,r,spec.get("style","context"))
        if not r.feature_count:raise RuntimeError("no features embedded")
        r.status="embedded"
    except Exception as exc:r.status="failed";r.errors.append(f"{type(exc).__name__}: {exc}");r.errors.append(traceback.format_exc(limit=6))
    return r

def esri_geom_to_shape(g:Mapping[str,Any])->Any|None:
    if not g:return None
    if "x" in g and "y" in g:return Point(float(g["x"]),float(g["y"]))
    if "points" in g:return MultiPoint(g["points"])
    if "paths" in g:return LineString(g["paths"][0]) if len(g["paths"])==1 else MultiLineString(g["paths"])
    if "rings" in g:
        rings=g["rings"]
        if not rings:return None
        try:p=Polygon(rings[0],holes=rings[1:]);return p if p.is_valid else p.buffer(0)
        except Exception:
            polys=[Polygon(ring) for ring in rings if len(ring)>=4];return MultiPolygon(polys) if len(polys)>1 else (polys[0] if polys else None)
    return None

def arcgis_query(session:requests.Session,layer_url:str,headers:Mapping[str,str]|None=None,verify:bool=True,where:str="1=1")->tuple[gpd.GeoDataFrame,dict[str,Any]]:
    layer_url=layer_url.rstrip("/");h=dict(headers or {});meta_resp=session.get(layer_url,params={"f":"json"},headers=h,timeout=180,verify=verify);meta_resp.raise_for_status();meta=meta_resp.json()
    if meta.get("error"):raise RuntimeError(meta["error"])
    max_records=int(meta.get("maxRecordCount") or 1000);oid=meta.get("objectIdField") or meta.get("objectIdFieldName")
    if not oid:
        for f in meta.get("fields",[]):
            if f.get("type")=="esriFieldTypeOID":oid=f.get("name");break
    params_ids={"f":"json","where":where,"returnIdsOnly":"true","returnGeometry":"false"};ids_resp=session.post(layer_url+"/query",data=params_ids,headers=h,timeout=240,verify=verify)
    if ids_resp.status_code>=400:ids_resp=session.get(layer_url+"/query",params=params_ids,headers=h,timeout=240,verify=verify)
    ids_resp.raise_for_status();ids=ids_resp.json().get("objectIds");records=[];geoms=[]
    if ids is not None:
        ids=sorted(set(ids));batch_size=min(max_records,250)
        for i in range(0,len(ids),batch_size):
            batch=ids[i:i+batch_size];params={"f":"geojson","objectIds":",".join(map(str,batch)),"outFields":"*","returnGeometry":"true","outSR":"4326","geometryPrecision":"7"};resp=session.post(layer_url+"/query",data=params,headers=h,timeout=300,verify=verify)
            if resp.status_code<400:
                try:
                    obj=resp.json()
                    if obj.get("type")=="FeatureCollection":
                        for feat in obj.get("features",[]):
                            if feat.get("geometry"):records.append(feat.get("properties") or {});geoms.append(shape(feat["geometry"]))
                        continue
                except Exception:pass
            params["f"]="json";resp=session.post(layer_url+"/query",data=params,headers=h,timeout=300,verify=verify);resp.raise_for_status();obj=resp.json()
            if obj.get("error"):raise RuntimeError(obj["error"])
            for feat in obj.get("features",[]):
                geom=esri_geom_to_shape(feat.get("geometry") or {})
                if geom is not None:records.append(feat.get("attributes") or {});geoms.append(geom)
    else:
        offset=0
        while True:
            params={"f":"geojson","where":where,"outFields":"*","returnGeometry":"true","outSR":"4326","geometryPrecision":"7","resultOffset":offset,"resultRecordCount":min(max_records,1000)}
            if oid:params["orderByFields"]=oid
            resp=session.post(layer_url+"/query",data=params,headers=h,timeout=300,verify=verify);resp.raise_for_status();obj=resp.json();feats=obj.get("features") or []
            for feat in feats:
                if feat.get("geometry"):records.append(feat.get("properties") or {});geoms.append(shape(feat["geometry"]))
            offset+=len(feats)
            if not feats or not obj.get("exceededTransferLimit"):break
    return gpd.GeoDataFrame(records,geometry=geoms,crs=4326),meta

def process_arcgis(builder:KMLBuilder,session:requests.Session,spec:Mapping[str,Any])->SourceResult:
    r=SourceResult(**{k:spec[k] for k in ("source_id","region","country","title","category","authority","source_url","method")});r.retrieved_at=utcnow();r.route_version=spec.get("route_version","official current snapshot");r.coverage=spec.get("coverage","")
    try:
        gdf,meta=arcgis_query(session,r.source_url,spec.get("headers"),spec.get("verify",True),spec.get("where","1=1"));top=builder.folder(r.region,r.title,f"{r.authority}\n{r.source_url}\nRetrieved {r.retrieved_at}");r.notes.append(f"ArcGIS layer: {meta.get('name','')} / {meta.get('geometryType','')}");builder.add_geodataframe(top,gdf,r,spec.get("style","context"))
        if not r.feature_count:raise RuntimeError("no features embedded")
        raw=gdf.to_json().encode("utf-8");r.downloaded_bytes=len(raw);r.sha256=sha256_bytes(raw);r.status="embedded"
    except Exception as exc:r.status="failed";r.errors.append(f"{type(exc).__name__}: {exc}");r.errors.append(traceback.format_exc(limit=6))
    return r

def find_service_urls(obj:Any)->set[str]:
    urls=set()
    def walk(v):
        if isinstance(v,dict):
            for x in v.values():walk(x)
        elif isinstance(v,list):
            for x in v:walk(x)
        elif isinstance(v,str):
            for m in re.finditer(r"https?://[^\s\"'<>]+/(?:MapServer|FeatureServer)(?:/\d+)?",html.unescape(v).replace("\\/","/"),re.I):urls.add(m.group(0).rstrip("/"))
    walk(obj);return urls

def arcgis_item_layers(session:requests.Session,item_id:str,portal:str="https://www.arcgis.com")->list[str]:
    base=portal.rstrip("/");seen=set();services=set();queue=[item_id]
    while queue and len(seen)<80:
        iid=queue.pop(0)
        if iid in seen:continue
        seen.add(iid);meta=session.get(f"{base}/sharing/rest/content/items/{iid}",params={"f":"json"},timeout=120,verify=False).json()
        if meta.get("url") and re.search(r"/(?:MapServer|FeatureServer)",meta["url"],re.I):services.add(meta["url"].rstrip("/"))
        dr=session.get(f"{base}/sharing/rest/content/items/{iid}/data",params={"f":"json"},timeout=120,verify=False)
        if dr.status_code<400:
            try:data=dr.json()
            except Exception:data={}
            services|=find_service_urls(data)
            for x in re.findall(r"\b[0-9a-fA-F]{32}\b",json.dumps(data)):
                if x not in seen:queue.append(x)
    layers=[]
    for svc in sorted(services):
        if re.search(r"/(?:MapServer|FeatureServer)/\d+$",svc,re.I):layers.append(svc);continue
        try:
            meta=session.get(svc,params={"f":"json"},timeout=120,verify=False).json()
            for layer in meta.get("layers",[]):layers.append(f"{svc}/{layer['id']}")
        except Exception:pass
    return sorted(set(layers))

def process_arcgis_item(builder:KMLBuilder,session:requests.Session,spec:Mapping[str,Any])->SourceResult:
    r=SourceResult(**{k:spec[k] for k in ("source_id","region","country","title","category","authority","source_url","method")});r.retrieved_at=utcnow();r.route_version=spec.get("route_version","official current snapshot");r.coverage=spec.get("coverage","")
    try:
        layers=arcgis_item_layers(session,spec["item_id"],spec.get("portal","https://www.arcgis.com"))
        if not layers:raise RuntimeError("no public feature layers discovered")
        inc=re.compile(spec.get("include",".*"),re.I);top=builder.folder(r.region,r.title,f"{r.authority}\nArcGIS item {spec['item_id']}\nRetrieved {r.retrieved_at}")
        for layer in layers:
            try:
                meta=session.get(layer,params={"f":"json"},timeout=90,verify=False).json();lname=clean_text(meta.get("name") or layer)
                if not inc.search(lname):continue
                gdf,_=arcgis_query(session,layer,verify=False)
                if gdf.empty:continue
                lf=sub(top,"Folder");sub(lf,"name",lname);sub(lf,"visibility","0");builder.add_geodataframe(lf,gdf,r,spec.get("style","context"))
            except Exception as exc:r.notes.append(f"Layer failed {layer}: {exc}")
        if not r.feature_count:raise RuntimeError("discovered layers yielded no features")
        r.status="embedded";r.sha256=hashlib.sha256("\n".join(layers).encode()).hexdigest()
    except Exception as exc:r.status="failed";r.errors.append(f"{type(exc).__name__}: {exc}");r.errors.append(traceback.format_exc(limit=6))
    return r

DMS_RE=re.compile(r"(?P<deg>\d{1,3})(?:\s*[°º]|\s+)(?P<min>\d{1,2}(?:\.\d+)?)?(?:\s*[′'’]|\s+)?(?P<sec>\d{1,2}(?:\.\d+)?)?\s*[″\"”]?\s*(?P<hem>[NSEW])",re.I)
DEC_RE=re.compile(r"(?P<val>\d{1,3}(?:\.\d+)?)\s*[°º]?\s*(?P<hem>[NSEW])",re.I)
def parse_coord_tokens(line:str)->list[tuple[str,float,int]]:
    vals=[];occupied=[]
    for m in DMS_RE.finditer(line):
        val=float(m.group("deg"))+float(m.group("min") or 0)/60+float(m.group("sec") or 0)/3600;hem=m.group("hem").upper();val=-val if hem in "SW" else val;vals.append(("lat" if hem in "NS" else "lon",val,m.start()));occupied.append(m.span())
    for m in DEC_RE.finditer(line):
        if any(a<=m.start()<b for a,b in occupied):continue
        val=float(m.group("val"));hem=m.group("hem").upper();val=-val if hem in "SW" else val;vals.append(("lat" if hem in "NS" else "lon",val,m.start()))
    return sorted(vals,key=lambda x:x[2])
def coordinate_pairs_from_text(text:str)->list[tuple[float,float,str]]:
    pairs=[]
    for line in text.splitlines():
        vals=parse_coord_tokens(line);lat=lon=None
        for kind,value,_ in vals:
            if kind=="lat" and lat is None:lat=value
            elif kind=="lon" and lon is None:lon=value
            if lat is not None and lon is not None:
                if -90<=lat<=90 and -180<=lon<=180:pairs.append((lon,lat,clean_text(line.strip())[:500]))
                lat=lon=None
    out=[]
    for p in pairs:
        if not out or (round(out[-1][0],7),round(out[-1][1],7))!=(round(p[0],7),round(p[1],7)):out.append(p)
    return out
def pdf_to_coordinates(pdf:Path)->tuple[list[tuple[float,float,str]],str]:
    proc=subprocess.run(["pdftotext","-layout",str(pdf),"-"],capture_output=True,check=False);text=proc.stdout.decode("utf-8","replace");return coordinate_pairs_from_text(text),text
def process_notice_pdf(builder:KMLBuilder,session:requests.Session,*,source_id:str,region:str,country:str,title:str,authority:str,url:str,work:Path,style:str="permit_work")->SourceResult:
    r=SourceResult(source_id,region,country,title,"Cable permit / navigation notice coordinates",authority,url,"official_pdf_coordinate_table");r.retrieved_at=utcnow();r.route_version="permitted/planned works or repair notice";r.coverage="coordinate sequence contained in notice"
    try:
        data,_=download(session,url,work/f"{safe_id(source_id)}.pdf",timeout=240,verify=False);r.downloaded_bytes=len(data);r.sha256=sha256_bytes(data);coords,_=pdf_to_coordinates(work/f"{safe_id(source_id)}.pdf")
        if len(coords)<1:raise RuntimeError("no coordinate pairs extracted from PDF text")
        folder=builder.folder(region,title,f"{authority}\n{url}\nCoordinates extracted from the official notice table/text; original order preserved.");props={"document_coordinate_count":len(coords),"coordinate_source":"official PDF text/table"}
        if len(coords)==1:builder.add_manual(folder,Point(coords[0][0],coords[0][1]),props|{"source_row":coords[0][2]},r,style)
        else:
            builder.add_manual(folder,LineString([(x,y) for x,y,_ in coords]),props|{"first_row":coords[0][2],"last_row":coords[-1][2]},r,style)
            for i,(x,y,row) in enumerate(coords,1):builder.add_manual(folder,Point(x,y),{"position_index":i,"source_row":row},r,style)
        r.status="embedded"
    except Exception as exc:r.status="failed";r.errors.append(f"{type(exc).__name__}: {exc}")
    return r
def scrape_pdf_links(session:requests.Session,page_url:str,keywords:str=r"submarine|cable|fibre|fiber|海底|光缆|海缆")->list[tuple[str,str]]:
    resp=session.get(page_url,timeout=180,verify=False);resp.raise_for_status();soup=BeautifulSoup(resp.text,"html.parser");out=[];rx=re.compile(keywords,re.I)
    for a in soup.find_all("a",href=True):
        href=urllib.parse.urljoin(resp.url,a["href"]);label=" ".join(a.stripped_strings);context=" ".join(a.parent.stripped_strings) if a.parent else label
        if (href.lower().endswith(".pdf") or ".pdf?" in href.lower()) and rx.search(label+" "+context+" "+href):out.append((label or Path(urllib.parse.urlparse(href).path).name,href))
    return list(dict.fromkeys(out))
def process_xlsx_coordinates(builder:KMLBuilder,session:requests.Session,spec:Mapping[str,Any],work:Path)->SourceResult:
    r=SourceResult(**{k:spec[k] for k in ("source_id","region","country","title","category","authority","source_url","method")});r.retrieved_at=utcnow();r.route_version=spec.get("route_version","regulatory coordinate schedule");r.coverage="coordinate tables"
    try:
        data,_=download(session,r.source_url,work/f"{r.source_id}.xlsx",verify=False);r.downloaded_bytes=len(data);r.sha256=sha256_bytes(data);wb=load_workbook(work/f"{r.source_id}.xlsx",read_only=True,data_only=True);root=builder.folder(r.region,r.title,f"{r.authority}\n{r.source_url}\nExact coordinate schedule converted from XLSX");total=0
        for ws in wb.worksheets:
            sequences=[];cur=[]
            for row in ws.iter_rows(values_only=True):
                line=" | ".join(clean_text(v) for v in row if v not in (None,""));pairs=coordinate_pairs_from_text(line)
                if pairs:cur.extend(pairs)
                elif cur:sequences.append(cur);cur=[]
            if cur:sequences.append(cur)
            for j,seq in enumerate(sequences,1):
                if not seq:continue
                lf=sub(root,"Folder");sub(lf,"name",f"{ws.title} — sequence {j}");sub(lf,"visibility","0")
                if len(seq)>=2:builder.add_manual(lf,LineString([(x,y) for x,y,_ in seq]),{"sheet":ws.title,"sequence":j,"position_count":len(seq)},r,spec.get("style","fish_closure"))
                for i,(x,y,line) in enumerate(seq,1):builder.add_manual(lf,Point(x,y),{"sheet":ws.title,"sequence":j,"position_index":i,"source_row":line},r,spec.get("style","fish_closure"))
                total+=len(seq)
        if not r.feature_count:raise RuntimeError("no coordinate pairs found")
        r.status="embedded";r.notes.append(f"Parsed {total} coordinate positions")
    except Exception as exc:r.status="failed";r.errors.append(f"{type(exc).__name__}: {exc}")
    return r
def process_ccamlr(builder:KMLBuilder,work:Path)->SourceResult:
    r=SourceResult("rfmo_ccamlr","International Waters","Southern Ocean","CCAMLR management and protected-area geometry","High-seas fisheries management / protected areas","Commission for the Conservation of Antarctic Marine Living Resources","https://github.com/ccamlr/data","official_git_geojson");r.retrieved_at=utcnow();r.route_version="current official repository snapshot";r.coverage="CCAMLR Convention Area management layers"
    try:
        repo=work/"ccamlr-data";subprocess.run(["git","clone","--depth","1","https://github.com/ccamlr/data.git",str(repo)],check=True,capture_output=True);root=builder.folder(r.region,r.title,f"Official CCAMLR Public Data Repository\n{r.source_url}");selected={"mpa","mpapd","oma","rb","ssru","ssmu","asd","ckfmu","sass"};files=[]
        for d in selected:files.extend((repo/"geographical_data"/d).glob("*.json"))
        hashes=[]
        for f in sorted(files):
            try:
                obj=json.loads(f.read_text(encoding="utf-8-sig"))
                if not isinstance(obj,dict) or obj.get("type") not in {"FeatureCollection","Feature"}:continue
                lf=sub(root,"Folder");sub(lf,"name",f.parent.name.upper()+" — "+f.stem);sub(lf,"visibility","0");builder.add_geojson(lf,obj,r,"marine_protected" if f.parent.name in {"mpa","mpapd","oma"} else "fish_management");hashes.append(sha256_bytes(f.read_bytes()))
            except Exception as exc:r.notes.append(f"Skipped {f}: {exc}")
        if not r.feature_count:raise RuntimeError("no CCAMLR GeoJSON features embedded")
        r.status="embedded";r.sha256=hashlib.sha256("".join(hashes).encode()).hexdigest()
    except Exception as exc:r.status="failed";r.errors.append(f"{type(exc).__name__}: {exc}")
    return r
def process_scraped_zip_page(builder:KMLBuilder,session:requests.Session,spec:Mapping[str,Any],work:Path)->list[SourceResult]:
    results=[]
    try:
        resp=session.get(spec["page"],timeout=180,verify=False);resp.raise_for_status();soup=BeautifulSoup(resp.text,"html.parser");links=[]
        for a in soup.find_all("a",href=True):
            href=urllib.parse.urljoin(resp.url,a["href"]);label=" ".join(a.stripped_strings)
            if re.search(r"\.zip(?:$|\?)",href,re.I) and re.search(spec.get("include",".*"),label+" "+href,re.I):links.append((label or Path(urllib.parse.urlparse(href).path).name,href))
        if not links:raise RuntimeError("no matching ZIP links on page")
        for i,(label,url) in enumerate(dict.fromkeys(links),1):
            child=dict(spec);child.update(source_id=f"{spec['source_id']}_{i}",title=f"{spec['title']} — {label}",source_url=url,method="official_download_zip");results.append(process_zip(builder,session,child,work))
    except Exception as exc:
        r=SourceResult(spec["source_id"],spec["region"],spec["country"],spec["title"],spec["category"],spec["authority"],spec["page"],"official_page_scrape");r.retrieved_at=utcnow();r.status="failed";r.errors.append(f"{type(exc).__name__}: {exc}");results.append(r)
    return results
def process_seafo(builder:KMLBuilder,session:requests.Session,work:Path)->list[SourceResult]:
    page="https://dataportal.saeri.org/dataset/existing-bottom-fishing-areas-and-areas-closed-for-fishing-within-the-limits-of-seafo-area";results=[]
    try:
        data=session.get("https://dataportal.saeri.org/api/3/action/package_search",params={"q":"SEAFO bottom fishing areas closed"},timeout=180,verify=False).json();resources=[]
        for pkg in data.get("result",{}).get("results",[]):
            for res in pkg.get("resources",[]):
                url=res.get("url","");fmt=(res.get("format") or "").lower()
                if fmt in {"zip","shp","geopackage","gpkg","geojson"} or re.search(r"\.(zip|gpkg|geojson)(?:$|\?)",url,re.I):resources.append((res.get("name") or "SEAFO vector",url))
        if not resources:raise RuntimeError("CKAN API returned no vector resources")
        for i,(label,url) in enumerate(resources,1):results.append(process_zip(builder,session,{"source_id":f"rfmo_seafo_{i}","region":"International Waters","country":"South-East Atlantic","title":f"SEAFO — {label}","category":"High-seas fishing footprint and closures","authority":"SEAFO measure geometry distributed by SAERI data portal","source_url":url,"method":"official_measure_derived_vector","style":"fish_closure","route_version":"CM 30/15 / portal snapshot","coverage":"SEAFO Convention Area"},work))
    except Exception as exc:
        r=SourceResult("rfmo_seafo","International Waters","South-East Atlantic","SEAFO existing bottom-fishing and closed areas","High-seas fishing footprint and closures","SEAFO measure geometry distributed by SAERI data portal",page,"CKAN_API");r.retrieved_at=utcnow();r.status="failed";r.errors.append(f"{type(exc).__name__}: {exc}");results.append(r)
    return results
def process_msil(builder:KMLBuilder,session:requests.Session)->list[SourceResult]:
    keys=["0e83ad5d93214e04abf37c970c32b641","10784fa6ea604de687b2052e55e03879","61b85294618247a6bf652a979c5a5bbc"];out=[]
    for sid,title,cat,url,style in [("jp_msil_submarine_cables","Japan MSIL submarine cable lines","Submarine cable routes","https://api.msil.go.jp/submarine-cable-line/v2/MapServer/2","cable_route")]:
        r=None
        for key in keys:
            r=process_arcgis(builder,session,{"source_id":sid,"region":"Asia / Pacific","country":"Japan","title":title,"category":cat,"authority":"Japan Coast Guard Hydrographic and Oceanographic Department — MSIL","source_url":url,"method":"official_api_static_snapshot","headers":{"Ocp-Apim-Subscription-Key":key},"verify":True,"style":style,"route_version":"electronic-chart published approximate cable position","coverage":"Japan waters"})
            if r.status=="embedded":break
        out.append(r)
    return out
def add_china_notices(builder:KMLBuilder)->list[SourceResult]:
    notices=[{"id":"cn_h14_repair_2023","title":"H14 submarine cable repair work corridor — 2023","url":"https://www.msa.gov.cn/html/hxaq/aqxx/hxtjg/GuangDong/20230713/1689229767302052983.html","coords":[(114+6/60+0.26/3600,22+10/60+19.70/3600),(114+6/60+2.78/3600,22+9/60+35.62/3600)],"kind":"line"},{"id":"cn_h4_repair_2024","title":"H4 submarine cable repair work area — 2024","url":"https://www.msa.gov.cn/html/hxaq/aqxx/hxtjg/GuangDong/20240207/1700432466314047000.html","coords":[(114+10/60+37.50/3600,22+6/60+48.20/3600)],"radius_nm":0.5,"kind":"circle"},{"id":"cn_h7_repair_2025","title":"H7 submarine cable repair work area — 2025","url":"https://www.msa.gov.cn/page/article.do?articleId=3A77395B-19E6-4747-A313-D9E6A2EDB6E8","coords":[(114+11/60+48.8/3600,22+7/60+32.3/3600)],"radius_nm":0.5,"kind":"circle"},{"id":"cn_h6_install_2026","title":"H6 submarine cable installation work area — 2026","url":"https://www.msa.gov.cn/html/hxaq/aqxx/hxtjg/GuangDong/20260401/18C3DCD0-D0D7-4D2B-BA25-D39DAD015D04.html","coords":[(114.1454,22.1673),(114.1728,22.1239),(114.2220,22.1049),(114.2508,22.1155)],"kind":"polygon"},{"id":"cn_h8_install_2026","title":"H8 submarine cable installation work area — 2026","url":"https://www.msa.gov.cn/html/hxaq/aqxx/hxtjg/GuangDong/20260327/B47CB101-D09F-4331-A83F-12A54C42A887.html","coords":[(114.1807,22.1430),(114.2132,22.1101),(114.2570,22.1015),(114.2774,22.1286)],"kind":"polygon"}];out=[]
    for n in notices:
        r=SourceResult(n["id"],"Asia / Pacific","China","China MSA — "+n["title"],"Submarine cable construction/repair permit area","Maritime Safety Administration of the People’s Republic of China",n["url"],"official_html_coordinate_notice");r.retrieved_at=utcnow();r.route_version="temporary permitted work/repair geometry";r.coverage="official navigation warning work area";r.sha256=hashlib.sha256(json.dumps(n,sort_keys=True).encode()).hexdigest()
        try:
            f=builder.folder(r.region,r.title,f"Historical/time-limited official navigation-warning geometry\n{r.source_url}")
            if n["kind"]=="line":geom=LineString(n["coords"])
            elif n["kind"]=="polygon":geom=Polygon(n["coords"]+[n["coords"][0]])
            else:
                lon,lat=n["coords"][0];radius_m=n["radius_nm"]*1852;from pyproj import Transformer;local=CRS.from_proj4(f"+proj=aeqd +lat_0={lat} +lon_0={lon} +datum=WGS84 +units=m +no_defs");fwd=Transformer.from_crs(4326,local,always_xy=True).transform;inv=Transformer.from_crs(local,4326,always_xy=True).transform;geom=transform(inv,transform(fwd,Point(lon,lat)).buffer(radius_m,resolution=64))
            builder.add_manual(f,geom,{"notice_type":"temporary cable works","source_coordinate_basis":"official MSA notice"},r,"permit_work");r.status="embedded"
        except Exception as exc:r.status="failed";r.errors.append(str(exc))
        out.append(r)
    return out
def validate(path:Path)->dict[str,Any]:
    tree=etree.parse(str(path),etree.XMLParser(huge_tree=True,resolve_entities=False,no_network=True));root=tree.getroot();links=root.xpath(".//*[local-name()='NetworkLink']");external=[e.text.strip() for e in root.xpath(".//*[local-name()='href']") if e.text and re.match(r"https?://",e.text.strip(),re.I)];placemarks=len(root.xpath(".//*[local-name()='Placemark']"));ids=[e.get("id") for e in root.xpath(".//*[@id]")];dup=sorted(k for k,v in __import__('collections').Counter(ids).items() if v>1);bad=0;vertices=0;bounds=[180,90,-180,-90]
    for e in root.xpath(".//*[local-name()='coordinates']"):
        for tok in re.split(r"\s+",(e.text or "").strip()):
            if not tok:continue
            try:
                a=tok.split(',');x=float(a[0]);y=float(a[1]);vertices+=1
                if not (-180.000001<=x<=180.000001 and -90.000001<=y<=90.000001):bad+=1
                else:bounds=[min(bounds[0],x),min(bounds[1],y),max(bounds[2],x),max(bounds[3],y)]
            except Exception:bad+=1
    return {"xml_parse_ok":True,"placemarks":placemarks,"coordinate_vertices":vertices,"invalid_coordinate_tokens":bad,"coordinates_valid":bad==0,"networklinks":len(links),"networklinks_zero":len(links)==0,"external_hrefs":external,"external_hrefs_zero":len(external)==0,"duplicate_ids":dup,"duplicate_ids_zero":not dup,"bounds":bounds if vertices else None,"sha256":hashlib.sha256(path.read_bytes()).hexdigest(),"bytes":path.stat().st_size}

def main(argv:Sequence[str]|None=None)->int:
    ap=argparse.ArgumentParser();ap.add_argument("--out",type=Path,default=Path("dist"));args=ap.parse_args(argv);args.out.mkdir(parents=True,exist_ok=True);work=args.out/"source_cache";work.mkdir(exist_ok=True);logging.basicConfig(level=logging.INFO,format="%(asctime)s %(levelname)s %(message)s");session=SessionFactory.make();builder=KMLBuilder("Global Fisheries & Offshore Infrastructure — Deep Permit and International-Waters Supplement");results=[]
    zip_specs=[
{"source_id":"au_acma_wa_cable_protection","region":"Asia / Pacific","country":"Australia","title":"Western Australia submarine cable protection zones and nominal routes","category":"Submarine cable routes and protection zones","authority":"Australian Communications and Media Authority","source_url":"https://www.acma.gov.au/sites/default/files/2019-10/Submarine%20Telecommunications%20Cables%20WA%20Protection%20Zone%20Geographic%20Coordinates.ZIP","method":"official_download_zip","style":"cable_zone","route_version":"declared protection-zone geometry","coverage":"Western Australia"},
{"source_id":"au_acma_nsw_cable_protection","region":"Asia / Pacific","country":"Australia","title":"New South Wales submarine cable protection zones and nominal routes","category":"Submarine cable routes and protection zones","authority":"Australian Communications and Media Authority","source_url":"https://www.acma.gov.au/sites/default/files/2020-04/submarine-cable-protection-zone-geographic-coordinates.zip","method":"official_download_zip","style":"cable_zone","route_version":"declared protection-zone geometry","coverage":"New South Wales"},
{"source_id":"au_acma_sspz_2025_variation","region":"Asia / Pacific","country":"Australia","title":"Southern Sydney protection-zone variation for Tabua cable — 2025","category":"Cable permit / protection-zone amendment","authority":"Australian Communications and Media Authority","source_url":"https://www.acma.gov.au/sites/default/files/2025-03/SSPZ%20shapefile%20-%20proposed%20variation.zip","method":"official_consultation_shapefile","style":"permit_work","route_version":"proposed/decided 2025 zone variation","coverage":"Southern Sydney"},
{"source_id":"nafo_divisions","region":"International Waters","country":"North-West Atlantic","title":"NAFO Divisions","category":"RFMO management divisions","authority":"Northwest Atlantic Fisheries Organization","source_url":"https://www.nafo.int/Portals/0/GIS/Divisions.zip","method":"official_download_zip","style":"fish_management","coverage":"NAFO Convention Area"},
{"source_id":"nafo_fishing_footprint","region":"International Waters","country":"North-West Atlantic","title":"NAFO bottom-fishing footprint","category":"High-seas bottom-fishing footprint","authority":"Northwest Atlantic Fisheries Organization","source_url":"https://www.nafo.int/Portals/0/GIS/FishingFootprint.zip","method":"official_download_zip","style":"fish_management","coverage":"NAFO Regulatory Area"},
{"source_id":"nafo_vme_closures","region":"International Waters","country":"North-West Atlantic","title":"NAFO VME closures","category":"High-seas fisheries closures","authority":"Northwest Atlantic Fisheries Organization","source_url":"https://www.nafo.int/Portals/0/GIS/VME_Closures.zip","method":"official_download_zip","style":"fish_closure","coverage":"NAFO Regulatory Area"},
{"source_id":"nafo_seamount_closures","region":"International Waters","country":"North-West Atlantic","title":"NAFO seamount closures","category":"High-seas fisheries closures","authority":"Northwest Atlantic Fisheries Organization","source_url":"https://www.nafo.int/Portals/0/GIS/Seamount_Closures.zip","method":"official_download_zip","style":"fish_closure","coverage":"NAFO Regulatory Area"},
{"source_id":"neafc_midatlantic_vme","region":"International Waters","country":"North-East Atlantic","title":"NEAFC Mid-Atlantic VME closures","category":"High-seas fisheries closures","authority":"North-East Atlantic Fisheries Commission","source_url":"https://www.neafc.org/system/files/mid-atlantic-vme-closures.zip","method":"official_download_zip","style":"fish_closure","coverage":"NEAFC Regulatory Area"},
{"source_id":"neafc_hatton_rockall","region":"International Waters","country":"North-East Atlantic","title":"NEAFC Hatton-Rockall Basin closures","category":"High-seas fisheries closures","authority":"North-East Atlantic Fisheries Commission","source_url":"https://www.neafc.org/system/files/hatton-rockall-basin-closures.zip","method":"official_download_zip","style":"fish_closure","coverage":"NEAFC Regulatory Area"},
{"source_id":"neafc_irminger","region":"International Waters","country":"North-East Atlantic","title":"NEAFC Irminger Sea closures","category":"High-seas fisheries closures","authority":"North-East Atlantic Fisheries Commission","source_url":"https://www.neafc.org/system/files/irminger-sea-closures.zip","method":"official_download_zip","style":"fish_closure","coverage":"NEAFC Regulatory Area"},
{"source_id":"sprfmo_bottom_management","region":"International Waters","country":"South Pacific","title":"SPRFMO bottom-fishing management areas","category":"High-seas bottom-fishing management areas","authority":"South Pacific Regional Fisheries Management Organisation","source_url":"https://www.sprfmo.int/assets/0-2023-Annual-Meeting/Meeting-Reports/SC11/Other/SC11-Bottom-Fishing-management-areas.zip","method":"official_download_zip","style":"fish_closure","coverage":"SPRFMO Convention Area"}]
    for spec in zip_specs:results.append(process_zip(builder,session,spec,work))
    for sid,title,url in [("neafc_rec_19_2014","NEAFC Recommendation 19:2014 coordinate schedule","https://www.neafc.org/system/files/rec-19-2014-xlsx-0.xlsx"),("neafc_rec_10_2018","NEAFC Recommendation 10:2018 coordinate schedule","https://www.neafc.org/system/files/rec-10-2018-xlsx.xlsx"),("neafc_rec_17_2019","NEAFC Recommendation 17:2019 coordinate schedule","https://www.neafc.org/system/files/rec-17-2019-xlsx.xlsx"),("neafc_rec_19_2019","NEAFC Recommendation 19:2019 coordinate schedule","https://www.neafc.org/system/files/rec-19-2019-xlsx.xlsx")]:results.append(process_xlsx_coordinates(builder,session,{"source_id":sid,"region":"International Waters","country":"North-East Atlantic","title":title,"category":"High-seas closure coordinate schedule","authority":"North-East Atlantic Fisheries Commission","source_url":url,"method":"official_xlsx_coordinate_table","style":"fish_closure"},work))
    results.append(process_ccamlr(builder,work));results.extend(process_scraped_zip_page(builder,session,{"source_id":"rfmo_siofa","region":"International Waters","country":"Southern Indian Ocean","title":"SIOFA protected and management areas","category":"High-seas protected / managed areas","authority":"Southern Indian Ocean Fisheries Agreement","page":"https://siofa.org/management/pa","style":"fish_closure","include":r"shapefile|shape|protected|management"},work));results.extend(process_seafo(builder,session,work))
    arc_specs=[
{"source_id":"cl_amerb_decreed","region":"South America","country":"Chile","title":"Decreed AMERB benthic management areas","category":"Legally decreed fisheries management areas","authority":"Chile SUBPESCA","source_url":"https://geoportal.subpesca.cl/server/rest/services/IDE_PUBLICO/SRMPUB_AMERB/MapServer/0","method":"official_arcgis_post_snapshot","style":"fish_management","coverage":"Chile"},
{"source_id":"cl_amerb_requested","region":"South America","country":"Chile","title":"Requested AMERB areas","category":"Requested fisheries management areas","authority":"Chile SUBPESCA","source_url":"https://geoportal.subpesca.cl/server/rest/services/IDE_PUBLICO/SRMPUB_AMERB/MapServer/1","method":"official_arcgis_post_snapshot","style":"permit_work","coverage":"Chile"},
{"source_id":"cl_amerb_rejected","region":"South America","country":"Chile","title":"Rejected AMERB applications","category":"Rejected/historic fisheries management applications","authority":"Chile SUBPESCA","source_url":"https://geoportal.subpesca.cl/server/rest/services/IDE_PUBLICO/SRMPUB_AMERB/MapServer/2","method":"official_arcgis_post_snapshot","style":"context","coverage":"Chile"},
{"source_id":"za_petroleum_exploitation_wells","region":"Africa","country":"South Africa","title":"Petroleum exploitation wellheads","category":"Offshore petroleum wellheads","authority":"South Africa DFFE marine spatial planning service","source_url":"https://screening.environment.gov.za/server/rest/services/OC_MSP/3_11_NDIR_Minerals/MapServer/2","method":"official_arcgis_post_snapshot","style":"infrastructure","coverage":"South Africa"},
{"source_id":"za_petroleum_exploration_wells","region":"Africa","country":"South Africa","title":"Petroleum exploration wellheads","category":"Offshore petroleum wellheads","authority":"South Africa DFFE marine spatial planning service","source_url":"https://screening.environment.gov.za/server/rest/services/OC_MSP/3_11_NDIR_Minerals/MapServer/3","method":"official_arcgis_post_snapshot","style":"infrastructure","coverage":"South Africa"},
{"source_id":"nz_benthic_protection","region":"Asia / Pacific","country":"New Zealand","title":"Benthic Protection Areas","category":"Commercial fishing restrictions","authority":"New Zealand Ministry for Primary Industries","source_url":"https://maps.mpi.govt.nz/wss/service/arcgis1/guest/MARINE/MARINE_Restrictions_CommercialFishingRegulations/MapServer/4","method":"official_arcgis_post_snapshot","style":"fish_closure","verify":False},
{"source_id":"nz_closed_seamounts","region":"Asia / Pacific","country":"New Zealand","title":"Closed Seamount Areas","category":"Commercial fishing restrictions","authority":"New Zealand Ministry for Primary Industries","source_url":"https://maps.mpi.govt.nz/wss/service/arcgis1/guest/MARINE/MARINE_Restrictions_CommercialFishingRegulations/MapServer/5","method":"official_arcgis_post_snapshot","style":"fish_closure","verify":False},
{"source_id":"nz_cable_pipeline_protection","region":"Asia / Pacific","country":"New Zealand","title":"Submarine Cables and Pipelines Protection Areas","category":"Cable and pipeline protection areas","authority":"New Zealand Ministry for Primary Industries","source_url":"https://maps.mpi.govt.nz/wss/service/arcgis1/guest/MARINE/MARINE_Restrictions_CommercialFishingRegulations/MapServer/16","method":"official_arcgis_post_snapshot","style":"cable_zone","verify":False,"route_version":"statutory protection area"},
{"source_id":"nz_trawl_prohibition","region":"Asia / Pacific","country":"New Zealand","title":"Commercial trawl prohibition areas","category":"Commercial fishing restrictions","authority":"New Zealand Ministry for Primary Industries","source_url":"https://maps.mpi.govt.nz/wss/service/arcgis1/guest/MARINE/MARINE_Restrictions_CommercialFishingRegulations/MapServer/17","method":"official_arcgis_post_snapshot","style":"fish_closure","verify":False},
{"source_id":"nz_trawl_restriction","region":"Asia / Pacific","country":"New Zealand","title":"Commercial trawl restriction areas","category":"Commercial fishing restrictions","authority":"New Zealand Ministry for Primary Industries","source_url":"https://maps.mpi.govt.nz/wss/service/arcgis1/guest/MARINE/MARINE_Restrictions_CommercialFishingRegulations/MapServer/18","method":"official_arcgis_post_snapshot","style":"fish_closure","verify":False},
{"source_id":"nz_mataitai","region":"Asia / Pacific","country":"New Zealand","title":"Mātaitai reserves","category":"Customary fishing reserves","authority":"New Zealand Ministry for Primary Industries","source_url":"https://maps.mpi.govt.nz/wss/service/arcgis1/guest/MARINE/MARINE_RestrictionsCustomary/MapServer/4","method":"official_arcgis_post_snapshot","style":"marine_protected","verify":False}]
    for spec in arc_specs:results.append(process_arcgis(builder,session,spec))
    item_specs=[
{"source_id":"boem_export_cables_fixed","region":"North America","country":"United States","title":"BOEM offshore wind export cables — resolved feature service","category":"Offshore power cables","authority":"U.S. Bureau of Ocean Energy Management","source_url":"https://www.arcgis.com/home/item.html?id=d90e369be3fe49fea26cb745cbc0584b","method":"official_arcgis_item_resolution","item_id":"d90e369be3fe49fea26cb745cbc0584b","style":"cable_route","include":r"export|cable"},
{"source_id":"boem_substations_fixed","region":"North America","country":"United States","title":"BOEM offshore wind substations — resolved feature service","category":"Offshore substations","authority":"U.S. Bureau of Ocean Energy Management","source_url":"https://www.arcgis.com/home/item.html?id=1dc4178929e2447a82e369d219b8eb52","method":"official_arcgis_item_resolution","item_id":"1dc4178929e2447a82e369d219b8eb52","style":"infrastructure","include":r"substation|offshore"},
{"source_id":"gfcm_fra","region":"International Waters","country":"Mediterranean and Black Sea","title":"GFCM Fisheries Restricted Areas","category":"RFMO fisheries restricted areas","authority":"FAO General Fisheries Commission for the Mediterranean","source_url":"https://experience.arcgis.com/experience/232ba2e7562a49888e457022ad0edc8a","method":"official_arcgis_experience_resolution","item_id":"232ba2e7562a49888e457022ad0edc8a","style":"fish_closure","include":r"restrict|closure|fra|fish"},
{"source_id":"gfcm_national_closures","region":"International Waters","country":"Mediterranean and Black Sea","title":"GFCM national fisheries closures","category":"National and RFMO fisheries closures","authority":"FAO General Fisheries Commission for the Mediterranean","source_url":"https://experience.arcgis.com/experience/c03b310793974b93969f8dace5d8859c","method":"official_arcgis_experience_resolution","item_id":"c03b310793974b93969f8dace5d8859c","style":"fish_closure","include":r"closure|fish|restrict"},
{"source_id":"thailand_fisheries_portal","region":"Asia / Pacific","country":"Thailand","title":"Thailand Department of Fisheries public marine layers","category":"Fisheries management and marine-use areas","authority":"Thailand Department of Fisheries","source_url":"https://gisportal.fisheries.go.th/portal/apps/webappviewer/index.html?id=d2308e424dbf4b9ca6fe4758e395c31d","method":"official_arcgis_webapp_resolution","item_id":"d2308e424dbf4b9ca6fe4758e395c31d","portal":"https://gisportal.fisheries.go.th/portal","style":"fish_management","include":r"fish|fisher|marine|ทะเล|ประมง|เขต|ปิด|อนุรักษ์|เพาะเลี้ยง|ชายฝั่ง"},
{"source_id":"malaysia_hydro_portal","region":"Asia / Pacific","country":"Malaysia","title":"Malaysia National Hydrographic Centre public marine layers","category":"Hydrographic and marine infrastructure data","authority":"National Hydrographic Centre Malaysia","source_url":"https://geohub.hydro.gov.my/portal/home/item.html?id=7a66e1af30ff4682a4392e2777aba47f","method":"official_arcgis_item_resolution","item_id":"7a66e1af30ff4682a4392e2777aba47f","portal":"https://geohub.hydro.gov.my/portal","style":"infrastructure","include":r"cable|pipeline|marine|hydro|park|fish|submarine"}]
    for spec in item_specs:results.append(process_arcgis_item(builder,session,spec))
    results.extend(process_msil(builder,session));results.extend(add_china_notices(builder))
    for year in range(2018,2027):
        page=f"https://marine21.marine.gov.my/appl/jict032008/jlsm/service/notice/nm_notice{year}.html"
        try:links=scrape_pdf_links(session,page)
        except Exception:links=[]
        for i,(label,url) in enumerate(links,1):results.append(process_notice_pdf(builder,session,source_id=f"my_ntm_{year}_{i}",region="Asia / Pacific",country="Malaysia",title=f"Malaysia NTM {year} — {label}",authority="National Hydrographic Centre Malaysia",url=url,work=work))
    for j,page in enumerate(["https://www.mpa.gov.sg/media-centre/details/pm-no.-85-of-2026---submarine-cable-installation-works-off-changi","https://www.mpa.gov.sg/media-centre/details/pm-no.-119-of-2025---submarine-cable-laying-work-off-changi-east","https://www.mpa.gov.sg/media-centre/details/pm-no.-03-of-2026---submarine-cable-laying-work-off-changi-east"],1):
        try:links=scrape_pdf_links(session,page,keywords=r".*")
        except Exception:links=[]
        for i,(label,url) in enumerate(links,1):results.append(process_notice_pdf(builder,session,source_id=f"sg_pmn_{j}_{i}",region="Asia / Pacific",country="Singapore",title=f"Singapore MPA cable works — {label}",authority="Maritime and Port Authority of Singapore",url=url,work=work))
    hk_page="https://www.epd.gov.hk/eia/files/applications/en/pp_937/aep_6064/progress/action_180045/ep_661/EP_661_2025.html"
    try:hk_links=scrape_pdf_links(session,hk_page,keywords=r".*")
    except Exception:hk_links=[]
    for i,(label,url) in enumerate(hk_links,1):results.append(process_notice_pdf(builder,session,source_id=f"hk_seah2x_{i}",region="Asia / Pacific",country="Hong Kong",title=f"SEA-H2X EP-661/2025 / PP-680/2025 — {label}",authority="Hong Kong Environmental Protection Department",url=url,work=work))
    kml=args.out/"global_deep_supplement.kml";kmz=args.out/"global_deep_supplement.kmz";builder.write(kml,kmz);val=validate(kml);summary={"generated_at":utcnow(),"sources_total":len(results),"sources_embedded":sum(r.status=="embedded" for r in results),"sources_failed":sum(r.status!="embedded" for r in results),"features_embedded":sum(r.feature_count for r in results),"vertices_embedded":sum(r.vertex_count for r in results),"validation":val,"sources":[asdict(r) for r in results]};(args.out/"global_deep_supplement_ledger.json").write_text(json.dumps(summary,ensure_ascii=False,indent=2),encoding="utf-8");fields=list(SourceResult.__dataclass_fields__)
    with (args.out/"global_deep_supplement_ledger.csv").open("w",newline="",encoding="utf-8") as fh:
        w=csv.DictWriter(fh,fieldnames=fields);w.writeheader()
        for r in results:
            row=asdict(r)
            for k,v in row.items():
                if isinstance(v,(list,dict)):row[k]=json.dumps(v,ensure_ascii=False,sort_keys=True)
            w.writerow(row)
    (args.out/"global_deep_supplement_validation.json").write_text(json.dumps(val,indent=2),encoding="utf-8");lines=["GLOBAL DEEP PERMIT / INTERNATIONAL-WATERS SUPPLEMENT",f"Generated: {summary['generated_at']}",f"Sources embedded: {summary['sources_embedded']} / {summary['sources_total']}",f"Features embedded: {summary['features_embedded']}",f"Coordinate vertices: {summary['vertices_embedded']}",f"NetworkLinks: {val['networklinks']}","","Failed sources:"]
    for r in results:
        if r.status!="embedded":lines.append(f"- {r.source_id}: {r.title}: {r.errors[0] if r.errors else 'no features'}")
    (args.out/"global_deep_supplement_summary.txt").write_text("\n".join(lines)+"\n",encoding="utf-8");shutil.rmtree(work,ignore_errors=True)
    return 0 if (val["xml_parse_ok"] and val["networklinks_zero"] and val["coordinates_valid"] and val["external_hrefs_zero"] and val["duplicate_ids_zero"] and val["placemarks"]>0) else 2
if __name__=="__main__":raise SystemExit(main())
